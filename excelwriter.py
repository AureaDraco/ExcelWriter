from re import match

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string


class ExcelSheet:

    def __init__(self, workbook, title="Sheet1", add_new_sheet=False):
        if not isinstance(title, str):
            raise TypeError("Title must be a string")

        if add_new_sheet:
            self.sheet = workbook.create_sheet(title=title)
        else:
            self.sheet = workbook.active
            self.sheet.title = title


class ExcelWorkbook:

    def __init__(self, filename='./Book1.xlsx', title="Sheet1", author=None):

        if not isinstance(filename, str):
            raise TypeError("File Name must be a string")

        self.wb = Workbook()

        if author:
            if not isinstance(author, str):
                raise TypeError("Author Name must be a string")
            self.wb.properties.creator = author

        self.filename = filename
        self.current_sheet = ExcelSheet(self.wb, title).sheet
        self.sheets = {title}
        self.num_sheets = 1

    @staticmethod
    def format_cell(cell_, format_param_dict):
        """
        Formats the given cell, based on the input parameters.

        :param cell_: Cell to be formatted
        :param format_param_dict: Dict of Cell Formatting Parameters
        :type format_param_dict: dict
        """

        font_size = format_param_dict.get("sz", 11)
        font_bold = format_param_dict.get("b", False)
        font_italics = format_param_dict.get("i", False)
        font_underline = format_param_dict.get("u", None)
        font_color = format_param_dict.get("color", "000000")
        cell_.font = Font(sz=font_size, b=font_bold, i=font_italics, u=font_underline, color=font_color)

        fill_type = format_param_dict.get("fill_type", None)
        fill_color = format_param_dict.get("fill", None)

        if fill_color:
            if fill_type:
                cell_.fill = PatternFill(patternType=fill_type, fgColor=fill_color)
            else:
                cell_.fill = PatternFill(patternType="solid", fgColor=fill_color)
        else:
            pass

        horizontal_alignment = format_param_dict.get("horizontal", "general")
        vertical_alignment = format_param_dict.get("vertical", "bottom")
        wrap_text = format_param_dict.get("wrap", False)
        cell_.alignment = Alignment(horizontal=horizontal_alignment, vertical=vertical_alignment, wrap_text=wrap_text)

    def format_row(self, row_index, formatting):
        """
        Formats the given row, based on the input parameters.

        :param row_index: Row Number of the Row to be formatted
        :type row_index: int
        :param formatting: Dict of Row Formatting Parameters
        :type formatting: dict
        """

        row = list(self.current_sheet.iter_rows())[row_index - 1]
        for cell_ in row:
            self.format_cell(cell_, formatting)

    def format_column(self, column_string, formatting):
        """
        Formats the given column, based on the input parameters.

        :param column_string: Column String of the Column to be formatted
        :type column_string: string
        :param formatting: Dict of Column Formatting Parameters
        :type formatting: dict
        """

        column_index = column_index_from_string(column_string)
        col = list(self.current_sheet.iter_cols())[column_index - 1]
        for cell_ in col:
            self.format_cell(cell_, formatting)

    def format_range(self, selection_range, formatting):
        """
        Formats the given range of selection, based on the input parameters.

        :param selection_range: Range of Cells to be formatted
        :type selection_range: str
        :param formatting: Dict of Formatting Parameters
        :type formatting: dict
        """

        start_column, start_row, end_column, end_row = get_range_coordinates(selection_range, True)
        for row_num in range(start_row, end_row + 1):
            for col_num in range(start_column, end_column + 1):
                cell_ = self.current_sheet[get_cell_string(col_num, row_num)]
                self.format_cell(cell_, formatting)

    def set_borders(self, selection_range, border_type='all', style='thin'):
        """
        Sets Borders to a given range of selection.

        :param selection_range: Range of Cells to be formatted
        :type selection_range: str
        :param border_type: All Borders (all); Outside Borders (outside)
        :type border_type: str
        :param style: Border-Style
        :type style: str
        """
        start_column, start_row, end_column, end_row = get_range_coordinates(selection_range, True)

        def all_borders(cell):
            border_ = Border(left=Side(style=style),
                             right=Side(style=style),
                             top=Side(style=style),
                             bottom=Side(style=style))
            cell.border = border_

        def outside_borders(cell):
            col_, row_ = get_cell_coordinates(cell.coordinate, numeric=True)

            # Set Border for Top Left Corner
            if (row_, col_) == (start_row, start_column):
                cell.border = Border(left=Side(style=style), top=Side(style=style))
            # Set Border for Top Right Corner
            elif (row_, col_) == (start_row, end_column):
                cell.border = Border(right=Side(style=style), top=Side(style=style))
            # Set Border for Bottom Left Corner
            elif (row_, col_) == (end_row, start_column):
                cell.border = Border(left=Side(style=style), bottom=Side(style=style))
            # Set Border for Bottom Right Corner
            elif (row_, col_) == (end_row, end_column):
                cell.border = Border(right=Side(style=style), bottom=Side(style=style))
            # Set Border for Top Row
            elif row_ == start_row:
                cell.border = Border(top=Side(style=style))
            # Set Border for Left Column
            elif col_ == start_column:
                cell.border = Border(left=Side(style=style))
            # Set Border for Bottom Row
            elif row_ == end_row:
                cell.border = Border(bottom=Side(style=style))
            # Set Border for Right Column
            elif col_ == end_column:
                cell.border = Border(right=Side(style=style))
            # Skip setting Borders for the Cells in the Middle
            else:
                pass

        dispatch_table = {'all': all_borders, 'outside': outside_borders}

        if border_type in dispatch_table:
            for row_num in range(start_row, end_row + 1):
                for col_num in range(start_column, end_column + 1):
                    cell_ = self.current_sheet[get_cell_string(col_num, row_num)]
                    dispatch_table[border_type](cell_)
        else:
            raise ValueError(f"'{border_type}' is not a valid border type.")

    def insert_data(self, data, row_offset=1, col_offset=1):
        """
        Inserts data into a Spreadsheet.

        :param data: List of Lists to inserted in the Spreadsheet
        :type data: list
        :param row_offset: Starting Row
        :type row_offset: int
        :param col_offset: Starting Column
        :type col_offset: int
        """

        for row_index, row in enumerate(data, row_offset):
            for col_index, cell_value in enumerate(row, col_offset):
                cell_ = get_cell_string(col_index, row_index)
                self.current_sheet[cell_] = cell_value

    def add_sheet(self, title):
        """
        Adds a new sheet with the given title name to the Spreadsheet

        :param title: Title of the Sheet to be added
        :type title: str
        """

        if title not in self.sheets:
            self.current_sheet = ExcelSheet(self.wb, title, True).sheet
            self.sheets.add(title)
            self.num_sheets += 1
        else:
            raise ValueError(f"Cannot Add New Sheet '{title}'. Sheet '{title}' already exists.")

    def remove_sheet(self, title):
        """
        Removes an existing sheet with the given title name from the Spreadsheet

        :param title: Title of the Sheet to be removed
        :type title: str
        """
        if title in self.sheets:
            del self.wb[title]
            self.sheets.remove(title)
            self.num_sheets -= 1
        else:
            raise ValueError(f"Cannot Remove Sheet '{title}'. Sheet '{title}' does not exist.")

    def rename_sheet(self, old_title, new_title):
        """
        Renames an existing sheet with the given title name from the Spreadsheet

        :param old_title: Title of the Sheet to be renamed
        :type old_title: str
        :param new_title: New title of the Sheet
        :type new_title: str
        """
        if old_title in self.sheets:

            if new_title not in self.sheets:
                self.wb[old_title].title = new_title
                self.sheets.remove(old_title)
                self.sheets.add(new_title)
            else:
                raise ValueError(f"Cannot Rename Sheet '{old_title}'. Sheet '{new_title}' already exists.")

        else:
            raise ValueError(f"Cannot Rename Sheet '{old_title}'. Sheet '{old_title}' does not exist.")

    def set_current_sheet(self, title):
        """
        Sets the Active Sheet to the Sheet with the given title.

        :param title: Title of the Sheet to be set active
        :type title: str
        """
        if title in self.sheets:
            self.current_sheet = self.wb[title]
        else:
            raise ValueError(f"Cannot set Current Sheet to '{title}'. Sheet '{title}' does not exist.")

    def save_file(self):
        self.wb.save(filename=self.filename)


def get_cell_string(col_num, row_num):
    """
    Returns the Cell String.
    
    :param col_num: Column Number
    :type col_num: int
    :param row_num: Row Number
    :type row_num: int
    :return: Cell Coordinates
    :rtype: str
    """

    return f"{get_column_letter(col_num)}{row_num}"  # .format(get_column_letter(col_num), row_num)


def get_cell_coordinates(cell_string, numeric=False):
    """

    :param cell_string:
    :type cell_string: str
    :param numeric:
    :type numeric: bool
    :return: 
    :rtype: tuple
    """

    m = match("([A-Z]+)([0-9]+)", cell_string.upper())

    if m:
        column_letter = m.group(1)
        row_number = int(m.group(2))
    else:
        raise ValueError(f"Cell Value: '{cell_string}' is not in the valid format")

    if numeric:
        return int(column_index_from_string(column_letter)), row_number
    else:
        return column_letter, row_number


def get_range_coordinates(cell_range, numeric=False):
    """
    Returns the Corner Cell Coordinates of a given Cell Range.

    :param cell_range: Cell Range
    :type cell_range: str
    :param numeric:
    :type numeric: bool
    :return:
    :rtype: tuple
    """

    start, end = cell_range.split(":")
    start_column, start_row = get_cell_coordinates(start, numeric)
    end_column, end_row = get_cell_coordinates(end, numeric)

    return start_column, start_row, end_column, end_row


if __name__ == "__main__":
    file_name = '/Users/animishr/Downloads/test.xlsx'
    sht = ExcelWorkbook(file_name, 'test', author='Animesh')
    sht.insert_data(["Hello", "World"])
    sht.format_column('A', {'sz': 12, 'b': True})
    sht.format_row(2, {'i': True})
    sht.format_range('B1:D2', {'color': "FF0000", 'fill': "00FF00"})
    sht.set_borders('B4:E5', 'all')
    # sht.add_sheet('test')
    print(sht.sheets, sht.num_sheets)
    sht.insert_data(["Hell", "yeah"])
    print(sht.sheets, sht.num_sheets)
    sht.set_current_sheet('test')
    sht.insert_data(["Hello", "World"], 4, 2)
    sht.rename_sheet('test', 'yoyo')
    print(sht.sheets, sht.num_sheets)
    sht.add_sheet('test')
    print(sht.sheets, sht.num_sheets)
    sht.remove_sheet('test')
    print(sht.sheets, sht.num_sheets)
    sht.save_file()
